"""
Writes the cleaned catalog out to a formatted, multi-tab Excel workbook:
  - Catalog: the clean product records
  - Data Quality Report: counts behind the cleaning run
  - Removed Duplicates: what got dropped and why, for audit purposes
  - Flagged Records: rows still missing a required field or category
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
FLAG_FILL = PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid")


def _write_dataframe(ws, df, freeze_header=True, highlight_col=None):
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    for _, row in df.iterrows():
        values = ["" if v is None else v for v in row.tolist()]
        ws.append(values)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    if highlight_col and highlight_col in df.columns:
        col_idx = list(df.columns).index(highlight_col) + 1
        for row_num in range(2, ws.max_row + 1):
            flag_cell = ws.cell(row=row_num, column=col_idx)
            if str(flag_cell.value).lower() in ("true", "1"):
                for cell in ws[row_num]:
                    cell.fill = FLAG_FILL

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    if freeze_header:
        ws.freeze_panes = "A2"


def export(clean_df, removed_dupes_df, quality_report_df, output_path=None):
    output_path = output_path or config.EXCEL_OUTPUT_PATH
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    ws_catalog = wb.active
    ws_catalog.title = "Catalog"
    _write_dataframe(ws_catalog, clean_df, highlight_col="missing_required_field")

    ws_quality = wb.create_sheet("Data Quality Report")
    _write_dataframe(ws_quality, quality_report_df)

    ws_dupes = wb.create_sheet("Removed Duplicates")
    if len(removed_dupes_df) > 0:
        _write_dataframe(ws_dupes, removed_dupes_df)
    else:
        ws_dupes.append(["No duplicate records found in this run"])

    flagged = clean_df[
        (clean_df["missing_required_field"]) | (~clean_df["category_valid"])
    ]
    ws_flagged = wb.create_sheet("Flagged Records")
    if len(flagged) > 0:
        _write_dataframe(ws_flagged, flagged, highlight_col="missing_required_field")
    else:
        ws_flagged.append(["No flagged records in this run"])

    wb.save(output_path)
    print(f"workbook written to {output_path}")
    return output_path
